# -*- coding: utf-8 -*-
"""
Created on Fri May 11 15:21:12 2018
This script is used to draw heatmap
@author: liuyang
"""

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from matplotlib import pyplot as plt
from matplotlib import cm 
from matplotlib import axes

def draw_heatmap(data,xlabels,ylabels,numtotal,numsum):

    cmap = cm.get_cmap('YlOrBr',100)

    figure=plt.figure(facecolor='w')

    ax=figure.add_subplot(1,1,1,position=[0,0,2,3])

    ax.set_yticks(range(len(ylabels)))

    ax.set_yticklabels(ylabels,fontsize=20)

    ax.set_xticks(range(len(xlabels)))

    ax.set_xticklabels(xlabels, rotation=90,fontsize=20)
    
    plt.xlabel('Region where premature death occurred',fontsize=24)
    plt.ylabel('Region where consumption occured',fontsize=24)

    vmax=data[0][0]

    vmin=data[0][0]

    for i in data:

        for j in i:

            if j>vmax:

                vmax=j

            if j<vmin:

                vmin=j

    map=ax.imshow(data,interpolation='nearest',cmap=cmap,aspect='auto',vmin=vmin,vmax=vmax)
    p=0
    for i in numtotal:
        plt.text(0+p, -0.6, int(i), ha='left', rotation=90, wrap=True,fontsize=20)  
        p=p+1
    p=0
    for i in numsum:
        plt.text(12.6, 0.1+p, int(i), ha='left', rotation=0, wrap=True,fontsize=20)  
        p=p+1

    for i in range(0,13):
        for j in range(0,13):
            if i==j:
                plt.text(-0.3+i, 0.1+j, data[j,i], ha='left', rotation=0, wrap=True,fontsize=20,color='white')
            else:
                plt.text(-0.3+i, 0.1+j, data[j,i], ha='left', rotation=0, wrap=True,fontsize=20)  

    plt.savefig('heatmap', dpi=300,bbox_inches='tight') 
    plt.show()
    
    
    
wb = load_workbook("./heatmap.xlsx")
sheet = wb.get_sheet_by_name("Agr")
i=j=0
a=np.zeros([13,13])
for row in sheet.iter_rows('B3:N15'):
    j=0
    for cell in row: 
        a[i,j] = float(cell.value)
        j=j+1
    i=i+1
    
numtotal = np.zeros(13)
for row in sheet.iter_rows('B16:N16'):
    j=0
    for cell in row: 
        numtotal[j] = cell.value
        j=j+1
        
numsum = np.zeros(13)
j=0
for row in sheet.iter_rows('O3:O15'):
    for cell in row: 
        numsum[j] = cell.value
        j=j+1

xlabels=['China','Rest of\n East Asia','India','Rest of\n Asia ',' Russia ','Western\n Europe','Eastern\n Europe','USA','Canada','Middle east and\n north Africa','Latin\n America ','Sub-Saharan\n Africa','Rest of\n the world']

ylabels=['China','Rest of\n East Asia','India','Rest of\n Asia ',' Russia ','Western\n Europe','Eastern\n Europe','USA','Canada','Middle east and\n north Africa','Latin\n America ','Sub-Saharan\n Africa','Rest of\n the world']

draw_heatmap(a,xlabels,ylabels,numtotal,numsum)
 